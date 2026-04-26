#!/usr/bin/env python3
"""Invoice Builder - Flask app for managing Amazon Business order invoices."""

import os
import io
import csv
import re
import sqlite3
import zipfile
import smtplib
import base64
import urllib.request
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, date
from copy import copy
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file, make_response
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Image as RLImage
from openpyxl.drawing.image import Image as XLImage

import json as _json

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get('DATA_DIR', BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)
LOGO_PATH = os.path.join(BASE_DIR, 'static', 'zero-logo.png')
DB_PATH = os.path.join(DATA_DIR, 'invoices.db')
TEMPLATE_PATH = os.path.join(BASE_DIR, 'invoice-template.xlsx')
CONFIG_PATH = os.path.join(DATA_DIR, 'config.json')

# ─── Config ─────────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {'accepted_pos': ['9999', '99999']}

def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH) as f:
            return _json.load(f)
    return dict(DEFAULT_CONFIG)

def save_config(cfg):
    with open(CONFIG_PATH, 'w') as f:
        _json.dump(cfg, f, indent=2)

def get_accepted_pos():
    return load_config().get('accepted_pos', DEFAULT_CONFIG['accepted_pos'])

def send_email_with_attachments(files: list, to_email: str, subject: str, body: str,
                                  zip_data: bytes = None, zip_filename: str = None):
    """Send email via Resend API (primary) with SMTP fallback.
    files: list of (filename, data_bytes, mime_type) tuples.
    """
    cfg = load_config()
    resend_key = cfg.get('resend_api_key', '')
    if resend_key:
        _send_via_resend(resend_key, cfg, files, to_email, subject, body, zip_data, zip_filename)
    else:
        _send_via_smtp(cfg, files, to_email, subject, body, zip_data, zip_filename)

def _send_via_resend(api_key: str, cfg: dict, files: list, to_email: str,
                     subject: str, body: str, zip_data=None, zip_filename=None):
    from_email = cfg.get('resend_from', 'Invoice Builder <onboarding@resend.dev>')
    attachments = []
    if zip_data and not files:
        attachments.append({
            'filename': zip_filename or 'invoice.zip',
            'content': base64.b64encode(zip_data).decode(),
        })
    else:
        for filename, data, _ in files:
            attachments.append({
                'filename': filename,
                'content': base64.b64encode(data).decode(),
            })
    payload = _json.dumps({
        'from': from_email,
        'to': [to_email],
        'subject': subject,
        'text': body,
        'attachments': attachments,
    }).encode()
    req = urllib.request.Request(
        'https://api.resend.com/emails',
        data=payload,
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        method='POST',
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        if resp.status not in (200, 201):
            raise ValueError(f"Resend error {resp.status}: {resp.read().decode()}")

def _send_via_smtp(cfg: dict, files: list, to_email: str, subject: str, body: str,
                   zip_data=None, zip_filename=None):
    smtp_cfg = cfg.get('smtp', {})
    smtp_server = smtp_cfg.get('server', 'smtp.gmail.com')
    smtp_port = int(smtp_cfg.get('port', 587))
    smtp_user = smtp_cfg.get('user', '')
    smtp_password = smtp_cfg.get('password', '')
    if not smtp_user or not smtp_password:
        raise ValueError("Credenciales SMTP no configuradas")
    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    if zip_data and not files:
        attachment = MIMEBase('application', 'zip')
        attachment.set_payload(zip_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=zip_filename)
        msg.attach(attachment)
    else:
        for filename, data, mime_type in files:
            maintype, subtype = mime_type.split('/', 1) if '/' in mime_type else ('application', 'octet-stream')
            attachment = MIMEBase(maintype, subtype)
            attachment.set_payload(data)
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', 'attachment', filename=filename)
            msg.attach(attachment)
    with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(smtp_user, smtp_password)
        server.sendmail(smtp_user, to_email, msg.as_string())

# ─── Database ───────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT NOT NULL,
            asin TEXT,
            title TEXT NOT NULL,
            price REAL NOT NULL,
            qty INTEGER NOT NULL,
            invoice_id INTEGER,
            status TEXT DEFAULT 'pending',
            po TEXT,
            order_date TEXT,
            order_status TEXT,
            total_neto REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (invoice_id) REFERENCES invoices(id)
        );
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number INTEGER NOT NULL UNIQUE,
            date TEXT NOT NULL,
            total REAL NOT NULL,
            items_count INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    # Migration: add type column if missing
    try:
        conn.execute("ALTER TABLE invoices ADD COLUMN type TEXT DEFAULT 'invoice'")
        conn.commit()
    except Exception:
        pass  # Column already exists
    # Migration: add tracking column if missing
    try:
        conn.execute("ALTER TABLE items ADD COLUMN tracking TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass  # Column already exists
    # Clean items with invalid qty
    conn.execute("DELETE FROM items WHERE qty <= 0 AND status = 'pending'")
    conn.commit()
    conn.close()

def get_next_invoice_number():
    conn = get_db()
    row = conn.execute("SELECT MAX(invoice_number) as max_num FROM invoices").fetchone()
    conn.close()
    db_max = row['max_num'] if row and row['max_num'] else 2069
    # Check config override
    cfg = load_config()
    config_max = cfg.get('last_invoice_number', 0)
    return max(db_max, config_max) + 1

# ─── CSV Parsing ────────────────────────────────────────────────────────────

def clean_csv_value(val):
    """Clean Amazon Business CSV values like =\"99999\" or \"12.99\""""
    if val is None:
        return ''
    val = val.strip()
    # Remove ="" wrapper
    if val.startswith('="') and val.endswith('"'):
        val = val[2:-1]
    # Remove surrounding quotes
    if val.startswith('"') and val.endswith('"'):
        val = val[1:-1]
    return val

def parse_price(val):
    """Parse price string to float."""
    val = clean_csv_value(val)
    val = val.replace(',', '')
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def parse_csv(file_content):
    """Parse Amazon Business CSV — only import items with 'Estado de entrega' == 'Entregado'."""
    # Try utf-8-sig first, then latin-1
    for encoding in ['utf-8-sig', 'utf-8', 'latin-1']:
        try:
            text = file_content.decode(encoding)
            break
        except (UnicodeDecodeError, AttributeError):
            if encoding == 'latin-1':
                text = file_content.decode('latin-1', errors='replace')

    reader = csv.DictReader(io.StringIO(text))
    items = []

    for row in reader:
        # FILTER 1: only accepted POs (configurable)
        po_raw = row.get('Número de PO', '')
        po_clean = clean_csv_value(po_raw)
        accepted_pos = get_accepted_pos()
        if po_clean not in accepted_pos:
            continue

        # FILTER 2: only "Entregado" delivery status
        estado_entrega = row.get('Estado de entrega', '').strip()
        if estado_entrega != 'Entregado':
            continue

        title = row.get('Cargo', '').strip()
        if not title:
            continue

        asin = row.get('ASIN', '').strip()
        order_id = row.get('Id. de pedido', '').strip()
        tracking = row.get('N.º de seguimiento del transportista', '').strip()
        qty_str = row.get('Cantidad de artículos', '1')
        price_str = row.get('Subtotal de artículo', '0')
        fecha = row.get('Fecha del pedido', '').strip()
        estado = row.get('Estado del pedido', '').strip()
        total_neto_str = row.get('Total neto del artículo', '0')

        qty = int(clean_csv_value(qty_str) or '1')
        if qty <= 0:
            continue
        price = parse_price(price_str)
        total_neto = parse_price(total_neto_str)

        # "Subtotal de artículo" is the TOTAL for all units — divide by qty to get unit price
        unit_price = round(price / qty, 2) if qty > 0 else price

        key = (order_id, asin, title)
        existing_idx = next((i for i, x in enumerate(items) if (x['order_id'], x['asin'], x['title']) == key), None)
        if existing_idx is not None:
            # Same item appears multiple times (Amazon splits units into separate rows) — merge
            prev = items[existing_idx]
            merged_qty = prev['qty'] + qty
            merged_total = round(prev['price'] * prev['qty'] + unit_price * qty, 2)
            items[existing_idx]['qty'] = merged_qty
            items[existing_idx]['price'] = round(merged_total / merged_qty, 2)
            items[existing_idx]['total_neto'] = round(prev['total_neto'] + total_neto, 2)
        else:
            items.append({
                'order_id': order_id,
                'asin': asin,
                'title': title,
                'price': unit_price,
                'qty': qty,
                'po': po_clean,
                'order_date': fecha,
                'order_status': estado,
                'total_neto': total_neto,
                'tracking': tracking,
            })

    return items

# ─── XLSX Generation ────────────────────────────────────────────────────────

def generate_xlsx(invoice_number, invoice_date, items):
    """Generate XLSX invoice matching the real Zero LLC template exactly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoice'

    # Column widths from template
    ws.column_dimensions['A'].width = 9.0
    ws.column_dimensions['B'].width = 40.0
    ws.column_dimensions['C'].width = 18.0
    ws.column_dimensions['D'].width = 10.38
    ws.column_dimensions['E'].width = 43.5
    ws.column_dimensions['F'].width = 13.88
    ws.column_dimensions['G'].width = 11.5
    ws.column_dimensions['H'].width = 9.5

    # Styles
    thin_border = Side(style='thin', color='FF000000')
    bold_11 = Font(name='Calibri', bold=True, size=11)
    bold_10 = Font(name='Calibri', bold=True, size=10)
    bold_9 = Font(name='Calibri', bold=True, size=9)
    bold_8 = Font(name='Calibri', bold=True, size=8)
    bold_7 = Font(name='Calibri', bold=True, size=7)
    normal_10 = Font(name='Calibri', size=10)
    normal_9 = Font(name='Calibri', size=9)
    normal_8 = Font(name='Calibri', size=8)
    normal_7 = Font(name='Calibri', size=7)
    center_align = Alignment(horizontal='center', vertical='center')
    center_wrap = Alignment(horizontal='center', vertical='center', wrapText=True)
    left_align = Alignment(horizontal='left')
    right_align = Alignment(horizontal='right')

    # Row 1 - Logo (image, no text)
    ws.row_dimensions[1].height = 60.75
    if os.path.exists(LOGO_PATH):
        logo = XLImage(LOGO_PATH)
        logo.width = 350
        logo.height = 124
        logo.anchor = 'A1'
        ws.add_image(logo, 'A1')

    # Row 2: Address line 1 + Invoice #
    ws['A2'] = '7265 NW 74TH ST UNIT 9\n'
    ws['A2'].font = bold_11
    ws['E2'] = 'Invoice #:'
    ws['E2'].font = bold_11
    ws['F2'] = invoice_number
    ws['F2'].font = Font(name='Calibri', bold=True, size=10)
    ws['F2'].alignment = right_align

    # Row 3: Address line 2 + Invoice Date
    ws.row_dimensions[3].height = 14.25
    ws['A3'] = 'MEDLEY, Florida 33166'
    ws['A3'].font = bold_10
    ws['E3'] = 'Invoice Date:'
    ws['E3'].font = bold_10
    ws['F3'] = invoice_date
    ws['F3'].font = Font(name='Arial', bold=True, size=10)
    ws['F3'].alignment = Alignment(horizontal='center')
    ws['F3'].number_format = 'MM/DD/YYYY'

    # Row 4: Customer
    ws.row_dimensions[4].height = 14.25
    ws['E4'] = 'Customer:'
    ws['E4'].font = bold_11
    ws['F4'] = 'Menta Granizada SRL'
    ws['F4'].font = bold_10

    # Row 5 - spacer
    ws.row_dimensions[5].height = 14.25

    # Row 6: Name (merged A6:C6)
    ws.row_dimensions[6].height = 14.25
    ws.merge_cells('A6:C6')
    ws['A6'] = 'Name: Menta Granizada SRL'
    ws['A6'].font = bold_10

    # Row 7: Address + Ship To
    ws.row_dimensions[7].height = 14.25
    ws['A7'] = 'Address:'
    ws['A7'].font = bold_10
    ws.merge_cells('B7:C7')
    ws['B7'] = 'Vera 1150'
    ws['B7'].font = normal_10
    ws['E7'] = 'Ship To:'
    ws['E7'].font = bold_10
    ws['F7'] = 'Same'
    ws['F7'].font = bold_10

    # Row 8: City, ZIP, Incoterms
    ws.row_dimensions[8].height = 14.25
    ws['A8'] = 'City:Buenos Aires,ARG'
    ws['A8'].font = bold_10
    ws['C8'] = 'ZIP: 1414'
    ws['C8'].font = normal_10
    ws['C8'].alignment = left_align
    ws['E8'] = 'Incoterms: EXW'
    ws['E8'].font = bold_9

    # Row 9: CUIT
    ws.row_dimensions[9].height = 14.25
    ws['A9'] = 'CUIT: 30-71670182-0'
    ws['A9'].font = bold_10

    # Row 10 - spacer
    ws.row_dimensions[10].height = 14.25

    # Row 11: Headers with blue fill
    ws.row_dimensions[11].height = 14.25
    header_fill = PatternFill(patternType='solid', fgColor='FF4472C4')
    header_font = Font(name='Calibri', size=9)
    headers = [
        ('A11', 'Qty'), ('B11', 'ITEM'), ('F11', 'Disc %'),
        ('G11', 'RATE $'), ('H11', 'Amount')
    ]
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.merge_cells('B11:E11')

    # Data rows (starting at row 12)
    item_title_font = Font(name='Calibri', bold=True, size=9)
    row_num = 12
    for item in items:
        ws.row_dimensions[row_num].height = 22.5

        # Qty
        cell_a = ws.cell(row=row_num, column=1, value=item['qty'])
        cell_a.font = normal_10
        cell_a.alignment = center_align

        # Title (merged B:E) — BOLD per template
        cell_b = ws.cell(row=row_num, column=2, value=item['title'])
        cell_b.font = item_title_font
        cell_b.alignment = center_wrap
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)

        # Rate
        cell_g = ws.cell(row=row_num, column=7, value=item['price'])
        cell_g.font = normal_10
        cell_g.alignment = center_align
        cell_g.number_format = '#,##0.00'

        # Amount formula
        cell_h = ws.cell(row=row_num, column=8)
        cell_h.value = f'=G{row_num}*A{row_num}'
        cell_h.font = normal_10
        cell_h.alignment = center_align
        cell_h.number_format = '#,##0.00'

        row_num += 1

    # Footer rows
    last_data_row = row_num - 1
    footer_start = last_data_row + 3  # 2 empty rows after last item

    # Terms of warranty
    ws.cell(row=footer_start, column=1, value='TERMS OF WARRANTY: 30 DAYS').font = bold_7
    ws.cell(row=footer_start + 1, column=1, value='NO CHARGER').font = bold_7

    # Bank info in column D
    ws.cell(row=footer_start + 1, column=4, value='Bank Acc. Name').font = normal_8
    ws.cell(row=footer_start + 2, column=4, value='Bank Acc. Nº').font = normal_8
    ws.cell(row=footer_start + 3, column=4, value='SWIFT:').font = normal_8

    # All sales in US Dollars
    ws.merge_cells(start_row=footer_start + 4, start_column=1, end_row=footer_start + 4, end_column=2)
    ws.cell(row=footer_start + 4, column=1, value='All sales are in US Dollars').font = bold_9

    # Totals area (footerStart + 8)
    total_row = footer_start + 8
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1, value='SIGNATURE: X___________________________________').font = normal_8

    ws.cell(row=total_row, column=5, value='Non Taxable Subtotal').font = normal_8
    ws.cell(row=total_row, column=8, value=f'=SUM(H12:H{last_data_row})').font = normal_10
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'

    ws.cell(row=total_row + 1, column=5, value='Taxable Subtotal').font = normal_8

    ws.cell(row=total_row + 2, column=5, value='Tax').font = normal_8

    ws.cell(row=total_row + 3, column=1, value='Customer Original').font = normal_7
    ws.cell(row=total_row + 3, column=5, value='Total Order').font = normal_8
    ws.cell(row=total_row + 3, column=8, value=f'=H{total_row}').font = normal_10
    ws.cell(row=total_row + 3, column=8).number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ─── PDF Generation ─────────────────────────────────────────────────────────

def generate_pdf(invoice_number, invoice_date, items):
    """Generate a PDF invoice matching the Zero LLC reference template exactly."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=letter,
        topMargin=0.4*inch, bottomMargin=0.4*inch,
        leftMargin=0.5*inch, rightMargin=0.5*inch,
    )
    PAGE_W = letter[0] - 1.0*inch  # usable width

    styles = getSampleStyleSheet()
    # Reusable paragraph styles
    bold11 = ParagraphStyle('B11', fontName='Helvetica-Bold', fontSize=11, leading=13)
    bold10 = ParagraphStyle('B10', fontName='Helvetica-Bold', fontSize=10, leading=12)
    bold9  = ParagraphStyle('B9',  fontName='Helvetica-Bold', fontSize=9, leading=11)
    bold8  = ParagraphStyle('B8',  fontName='Helvetica-Bold', fontSize=8, leading=10)
    bold7  = ParagraphStyle('B7',  fontName='Helvetica-Bold', fontSize=7, leading=9)
    norm10 = ParagraphStyle('N10', fontName='Helvetica', fontSize=10, leading=12)
    norm9  = ParagraphStyle('N9',  fontName='Helvetica', fontSize=9, leading=11)
    norm8  = ParagraphStyle('N8',  fontName='Helvetica', fontSize=8, leading=10)
    norm7  = ParagraphStyle('N7',  fontName='Helvetica', fontSize=7, leading=9)
    item_style = ParagraphStyle('ItemP', fontName='Helvetica-Bold', fontSize=9, leading=11,
                                 alignment=1)  # center

    elements = []

    # ── Logo ──
    if os.path.exists(LOGO_PATH):
        elements.append(RLImage(LOGO_PATH, width=2.2*inch, height=0.78*inch, hAlign='LEFT'))
        elements.append(Spacer(1, 6))

    # ── Header: address (left) + invoice info (right) in bordered table ──
    date_str = invoice_date.strftime('%-m/%-d/%Y') if isinstance(invoice_date, (date, datetime)) else str(invoice_date)
    header_data = [
        [Paragraph('7265 NW 74TH ST UNIT 9', bold11),
         Paragraph('Invoice #:', bold11),
         Paragraph(str(invoice_number), bold10)],
        [Paragraph('MEDLEY, Florida 33166', bold10),
         Paragraph('Invoice Date:', bold10),
         Paragraph(date_str, bold10)],
        ['',
         Paragraph('Customer:', bold11),
         Paragraph('Menta Granizada SRL', bold10)],
    ]
    left_w = PAGE_W * 0.45
    label_w = PAGE_W * 0.22
    val_w = PAGE_W * 0.33
    header_table = Table(header_data, colWidths=[left_w, label_w, val_w])
    header_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.75, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (0, -1), 6),
        ('LEFTPADDING', (1, 0), (1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 8))

    # ── Customer info section ──
    cust_data = [
        [Paragraph('Name: Menta Granizada SRL', bold10), '', '',
         '', ''],
        [Paragraph('Address:', bold10), Paragraph('Vera 1150', norm10), '',
         Paragraph('Ship To:', bold10), Paragraph('Same', bold10)],
        [Paragraph('City:Buenos Aires,ARG', bold10), '',
         Paragraph('ZIP: 1414', norm10),
         Paragraph('Incoterms: EXW', bold9), ''],
        [Paragraph('CUIT: 30-71670182-0', bold10), '', '', '', ''],
    ]
    cw = [PAGE_W*0.28, PAGE_W*0.17, PAGE_W*0.15, PAGE_W*0.22, PAGE_W*0.18]
    cust_table = Table(cust_data, colWidths=cw)
    cust_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('SPAN', (0, 0), (2, 0)),  # Name spans 3 cols
    ]))
    elements.append(cust_table)
    elements.append(Spacer(1, 10))

    # ── Items table: Qty | ITEM | Disc % | RATE $ | Amount ──
    blue_header = colors.Color(0.267, 0.447, 0.769)  # #4472C4
    table_data = [[
        Paragraph('Qty', ParagraphStyle('H', fontName='Helvetica', fontSize=9, textColor=colors.white, alignment=1)),
        Paragraph('ITEM', ParagraphStyle('H2', fontName='Helvetica', fontSize=9, textColor=colors.white, alignment=1)),
        Paragraph('Disc %', ParagraphStyle('H3', fontName='Helvetica', fontSize=9, textColor=colors.white, alignment=1)),
        Paragraph('RATE $', ParagraphStyle('H4', fontName='Helvetica', fontSize=9, textColor=colors.white, alignment=1)),
        Paragraph('Amount', ParagraphStyle('H5', fontName='Helvetica', fontSize=9, textColor=colors.white, alignment=1)),
    ]]

    subtotal = 0
    for item in items:
        amount = item['qty'] * item['price']
        subtotal += amount
        title_text = item['title'][:120] + ('...' if len(item['title']) > 120 else '')
        table_data.append([
            str(item['qty']),
            Paragraph(title_text, item_style),
            '',  # Disc %
            f"${item['price']:,.2f}",
            f"${amount:,.2f}",
        ])

    # Add empty rows to fill space (like reference)
    min_rows = max(0, 12 - len(items))
    for _ in range(min_rows):
        table_data.append(['', '', '', '', ''])

    icw = [0.5*inch, PAGE_W - 0.5*inch - 0.75*inch - 0.9*inch - 0.9*inch,
           0.75*inch, 0.9*inch, 0.9*inch]
    items_table = Table(table_data, colWidths=icw, repeatRows=1)
    items_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), blue_header),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data rows
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 10))

    # ── Footer section: 2 columns (left: warranty/signature, right: bank/totals) ──
    footer_data = [
        [Paragraph('TERMS OF WARRANTY: 30 DAYS', bold7), '',
         Paragraph('Bank Acc. Name', norm8), ''],
        [Paragraph('NO CHARGER', bold7), '',
         Paragraph('Bank Acc. Nº', norm8), ''],
        ['', '',
         Paragraph('SWIFT:', norm8), ''],
        [Paragraph('All sales are in US Dollars', bold9), '', '', ''],
        ['', '', '', ''],
        [Paragraph('SIGNATURE: X___________________________________', norm8), '',
         Paragraph('Non Taxable Subtotal', norm8),
         Paragraph(f"${subtotal:,.2f}", norm10)],
        ['', '',
         Paragraph('Taxable Subtotal', norm8), ''],
        ['', '',
         Paragraph('Tax', norm8), ''],
        [Paragraph('Customer Original', norm7), '',
         Paragraph('Total Order', norm8),
         Paragraph(f"${subtotal:,.2f}", norm10)],
    ]
    fw = [PAGE_W*0.38, PAGE_W*0.07, PAGE_W*0.30, PAGE_W*0.25]
    footer_table = Table(footer_data, colWidths=fw)
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        # Line above Total Order
        ('LINEABOVE', (2, -1), (3, -1), 0.75, colors.black),
        ('LINEBELOW', (2, -1), (3, -1), 1.5, colors.black),
        # Span signature across cols 0-1
        ('SPAN', (0, 5), (1, 5)),
        ('SPAN', (0, 8), (1, 8)),
        ('SPAN', (0, 3), (1, 3)),
    ]))
    elements.append(footer_table)

    doc.build(elements)
    output.seek(0)
    return output

# ─── Remito PDF Generation ──────────────────────────────────────────────────

def generate_remito_pdf(invoice_number, invoice_date, items):
    """Generate a simple remito PDF — austere, no logo, Helvetica."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=letter,
        topMargin=0.6*inch, bottomMargin=0.5*inch,
        leftMargin=0.7*inch, rightMargin=0.7*inch,
    )
    PAGE_W = letter[0] - 1.4*inch

    # Styles
    title_style = ParagraphStyle('RemitoTitle', fontName='Helvetica-Bold', fontSize=16, leading=20, alignment=1)
    bold10 = ParagraphStyle('RB10', fontName='Helvetica-Bold', fontSize=10, leading=13)
    norm10 = ParagraphStyle('RN10', fontName='Helvetica', fontSize=10, leading=13)
    norm9  = ParagraphStyle('RN9',  fontName='Helvetica', fontSize=9, leading=11)
    bold9  = ParagraphStyle('RB9',  fontName='Helvetica-Bold', fontSize=9, leading=11, alignment=1)
    norm8  = ParagraphStyle('RN8',  fontName='Helvetica', fontSize=8, leading=10)

    elements = []

    # Title
    elements.append(Paragraph('REMITO', title_style))
    elements.append(Spacer(1, 16))

    # Date and number
    date_str = invoice_date.strftime('%d/%m/%Y') if isinstance(invoice_date, (date, datetime)) else str(invoice_date)
    info_data = [
        [Paragraph(f'<b>Fecha:</b> {date_str}', norm10),
         Paragraph(f'<b>Remito #:</b> {invoice_number}', norm10)],
    ]
    info_table = Table(info_data, colWidths=[PAGE_W * 0.5, PAGE_W * 0.5])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    # Items table
    header_style = ParagraphStyle('RH', fontName='Helvetica-Bold', fontSize=9, leading=11, alignment=1)
    item_style = ParagraphStyle('RI', fontName='Helvetica', fontSize=9, leading=11, alignment=1)
    item_left = ParagraphStyle('RIL', fontName='Helvetica', fontSize=9, leading=11)

    table_data = [
        [Paragraph('Qty', header_style),
         Paragraph('Descripción', header_style),
         Paragraph('Precio', header_style),
         Paragraph('Total', header_style)],
    ]

    subtotal = 0
    for item in items:
        amount = item['qty'] * item['price']
        subtotal += amount
        title_text = item['title'][:100] + ('...' if len(item['title']) > 100 else '')
        table_data.append([
            Paragraph(str(item['qty']), item_style),
            Paragraph(title_text, item_left),
            Paragraph(f"${item['price']:,.2f}", item_style),
            Paragraph(f"${amount:,.2f}", item_style),
        ])

    col_widths = [0.5*inch, PAGE_W - 0.5*inch - 1.0*inch - 1.0*inch, 1.0*inch, 1.0*inch]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        # Header
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Grid — simple borders, no background
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 12))

    # Subtotal / Total
    totals_data = [
        ['', Paragraph('<b>Subtotal:</b>', norm10), Paragraph(f"${subtotal:,.2f}", norm10)],
        ['', Paragraph('<b>Total:</b>', norm10), Paragraph(f"${subtotal:,.2f}", norm10)],
    ]
    totals_table = Table(totals_data, colWidths=[PAGE_W - 2.0*inch, 1.0*inch, 1.0*inch])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LINEABOVE', (1, 1), (2, 1), 0.75, colors.black),
        ('LINEBELOW', (1, 1), (2, 1), 1.0, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 30))

    # Firma
    elements.append(Paragraph('___________________________', norm10))
    elements.append(Paragraph('Firma', norm8))

    doc.build(elements)
    output.seek(0)
    return output


def generate_remito_pdf_no_prices(invoice_number, invoice_date, items):
    """Generate a remito PDF WITHOUT prices — only Qty and Descripción."""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=letter,
        topMargin=0.6*inch, bottomMargin=0.5*inch,
        leftMargin=0.7*inch, rightMargin=0.7*inch,
    )
    PAGE_W = letter[0] - 1.4*inch

    # Styles
    title_style = ParagraphStyle('RemitoTitle', fontName='Helvetica-Bold', fontSize=16, leading=20, alignment=1)
    norm10 = ParagraphStyle('RN10', fontName='Helvetica', fontSize=10, leading=13)
    norm8  = ParagraphStyle('RN8',  fontName='Helvetica', fontSize=8, leading=10)

    elements = []

    # Title
    elements.append(Paragraph('REMITO', title_style))
    elements.append(Spacer(1, 16))

    # Date and number
    date_str = invoice_date.strftime('%d/%m/%Y') if isinstance(invoice_date, (date, datetime)) else str(invoice_date)
    info_data = [
        [Paragraph(f'<b>Fecha:</b> {date_str}', norm10),
         Paragraph(f'<b>Remito #:</b> {invoice_number}', norm10)],
    ]
    info_table = Table(info_data, colWidths=[PAGE_W * 0.5, PAGE_W * 0.5])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    # Items table — only Qty and Descripción
    header_style = ParagraphStyle('RH', fontName='Helvetica-Bold', fontSize=9, leading=11, alignment=1)
    item_style = ParagraphStyle('RI', fontName='Helvetica', fontSize=9, leading=11, alignment=1)
    item_left = ParagraphStyle('RIL', fontName='Helvetica', fontSize=9, leading=11)

    table_data = [
        [Paragraph('Qty', header_style),
         Paragraph('Descripción', header_style)],
    ]

    for item in items:
        title_text = item['title'][:100] + ('...' if len(item['title']) > 100 else '')
        table_data.append([
            Paragraph(str(item['qty']), item_style),
            Paragraph(title_text, item_left),
        ])

    col_widths = [0.5*inch, PAGE_W - 0.5*inch]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 30))

    # Firma
    elements.append(Paragraph('___________________________', norm10))
    elements.append(Paragraph('Firma', norm8))

    doc.build(elements)
    output.seek(0)
    return output

# ─── Routes ─────────────────────────────────────────────────────────────────

def is_mobile():
    ua = request.headers.get('User-Agent', '')
    return bool(re.search(r'Mobile|Android|iPhone|iPad|iPod|webOS|BlackBerry|Opera Mini|IEMobile', ua, re.I))

@app.route('/')
def index():
    if is_mobile():
        return render_template('mobile.html')
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_csv():
    """Upload and parse Amazon Business CSV."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'File must be CSV'}), 400

    content = file.read()
    items = parse_csv(content)

    if not items:
        return jsonify({'error': 'No items found with "Estado de entrega" = Entregado'}), 400

    # Save to DB, skip duplicates
    conn = get_db()
    saved = 0
    skipped = 0
    for item in items:
        # Check if this order_id + asin combo already exists
        existing = conn.execute(
            "SELECT id FROM items WHERE order_id = ? AND asin = ? AND title = ?",
            (item['order_id'], item['asin'], item['title'])
        ).fetchone()
        if existing:
            skipped += 1
            continue

        conn.execute("""
            INSERT INTO items (order_id, asin, title, price, qty, po, order_date, order_status, total_neto, tracking, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending')
        """, (item['order_id'], item['asin'], item['title'], item['price'],
              item['qty'], item['po'], item['order_date'], item['order_status'],
              item['total_neto'], item['tracking']))
        saved += 1

    conn.commit()
    conn.close()

    return jsonify({
        'message': f'Se importaron {saved} ítems, se saltaron {skipped} duplicados',
        'saved': saved,
        'skipped': skipped,
        'total_parsed': len(items)
    })

@app.route('/api/items')
def get_items():
    """Get all items with their status."""
    conn = get_db()
    rows = conn.execute("""
        SELECT i.*, inv.invoice_number
        FROM items i
        LEFT JOIN invoices inv ON i.invoice_id = inv.id
        ORDER BY i.created_at DESC
    """).fetchall()

    # Count items per order_id for badge
    order_counts = {}
    for row in rows:
        oid = row['order_id']
        order_counts[oid] = order_counts.get(oid, 0) + 1

    items = []
    for row in rows:
        items.append({
            'id': row['id'],
            'order_id': row['order_id'],
            'asin': row['asin'],
            'title': row['title'],
            'price': row['price'],
            'qty': row['qty'],
            'po': row['po'],
            'order_date': row['order_date'],
            'order_status': row['order_status'],
            'total_neto': row['total_neto'],
            'tracking': row['tracking'],
            'status': row['status'],
            'invoice_id': row['invoice_id'],
            'invoice_number': row['invoice_number'],
            'multi_item': order_counts.get(row['order_id'], 0) > 1,
        })

    conn.close()
    return jsonify(items)

@app.route('/api/generate', methods=['POST'])
def generate_invoice():
    """Generate invoice from selected items."""
    data = request.json
    item_ids = data.get('items', [])

    if not item_ids:
        return jsonify({'error': 'No items selected'}), 400

    conn = get_db()

    # Fetch items
    placeholders = ','.join('?' * len(item_ids))
    rows = conn.execute(
        f"SELECT * FROM items WHERE id IN ({placeholders}) AND status = 'pending'",
        item_ids
    ).fetchall()

    if not rows:
        return jsonify({'error': 'No pending items found with given IDs'}), 400

    items = []
    order_ids = set()
    for row in rows:
        # Use quantity from request if provided
        qty_override = data.get('quantities', {}).get(str(row['id']))
        qty = int(qty_override) if qty_override else row['qty']
        items.append({
            'title': row['title'],
            'price': row['price'],
            'qty': qty,
            'order_id': row['order_id'],
            'tracking': row['tracking'] if 'tracking' in row.keys() else '',
        })
        order_ids.add(row['order_id'])

    # Apply discount if provided (silently — just use the reduced price)
    discount = float(data.get('discount', 0) or 0)
    if discount > 0:
        for item in items:
            item['price'] = round(item['price'] * (1 - discount / 100), 1)

    invoice_number = get_next_invoice_number()
    invoice_date = date.today()
    total = sum(i['price'] * i['qty'] for i in items)

    # Create invoice record
    cursor = conn.execute(
        "INSERT INTO invoices (invoice_number, date, total, items_count, type) VALUES (?, ?, ?, ?, ?)",
        (invoice_number, invoice_date.isoformat(), total, len(items), 'invoice')
    )
    invoice_id = cursor.lastrowid

    # Update items status
    conn.execute(
        f"UPDATE items SET status = 'invoiced', invoice_id = ? WHERE id IN ({placeholders})",
        [invoice_id] + item_ids
    )
    conn.commit()

    # Generate files
    xlsx_buf = generate_xlsx(invoice_number, invoice_date, items)
    pdf_buf = generate_pdf(invoice_number, invoice_date, items)
    remito_no_prices_buf = generate_remito_pdf_no_prices(invoice_number, invoice_date, items)

    # Generate TXT with trackings (grouped with item names)
    orders_grouped = defaultdict(list)
    for item in items:
        tracking = item.get('tracking', '') or item['order_id']
        orders_grouped[tracking].append(item['title'])
    lines = []
    for oid in sorted(orders_grouped.keys()):
        titles = orders_grouped[oid]
        if len(titles) > 1:
            truncated = [t[:60] for t in titles]
            lines.append(f"{oid} ({', '.join(truncated)})")
        else:
            lines.append(oid)
    txt_content = '\n'.join(lines)

    # Create ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'Invoice_{invoice_number}.xlsx', xlsx_buf.read())
        zf.writestr(f'Invoice_{invoice_number}.pdf', pdf_buf.read())
        zf.writestr(f'Remito_{invoice_number}-sin-precios.pdf', remito_no_prices_buf.read())
        zf.writestr(f'Invoice_{invoice_number}_trackings.txt', txt_content)

    zip_buf.seek(0)
    zip_data = zip_buf.read()
    conn.close()

    # Try to send email if configured — individual files, no ZIP
    email_status = 'not_configured'
    email_address = ''
    email_error = ''
    cfg = load_config()
    to_email = cfg.get('email')
    if to_email:
        email_address = to_email
        try:
            subject = f"Invoice {invoice_number} — Zero"

            # Build body with trackings
            body_lines = [f"Invoice #{invoice_number} — {invoice_date.strftime('%d/%m/%Y')}"]
            body_lines.append("")
            body_lines.append("Trackings:")
            body_lines.append("-" * 40)
            for line in lines:
                body_lines.append(line)
            body_lines.append("")
            body_lines.append(f"Total: ${total:,.2f}")
            body = '\n'.join(body_lines)

            # Regenerate buffers for email (the previous ones were consumed by ZIP)
            xlsx_buf_email = generate_xlsx(invoice_number, invoice_date, items)
            pdf_buf_email = generate_pdf(invoice_number, invoice_date, items)
            remito_np_email = generate_remito_pdf_no_prices(invoice_number, invoice_date, items)

            email_files = [
                (f'Invoice_{invoice_number}.xlsx', xlsx_buf_email.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                (f'Invoice_{invoice_number}.pdf', pdf_buf_email.read(), 'application/pdf'),
                (f'Remito_{invoice_number}-sin-precios.pdf', remito_np_email.read(), 'application/pdf'),
                (f'Invoice_{invoice_number}_trackings.txt', txt_content.encode('utf-8'), 'text/plain'),
            ]
            send_email_with_attachments(email_files, to_email, subject, body)
            email_status = 'sent'
        except Exception as exc:
            email_status = 'error'
            email_error = str(exc)

    response = make_response(send_file(
        io.BytesIO(zip_data),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Invoice_{invoice_number}.zip'
    ))
    response.headers['X-Email-Status'] = email_status
    response.headers['X-Email-Address'] = email_address
    response.headers['X-Email-Error'] = email_error
    response.headers['Access-Control-Expose-Headers'] = 'X-Email-Status, X-Email-Address, X-Email-Error'
    return response

@app.route('/api/generate-remito', methods=['POST'])
def generate_remito():
    """Generate remito from selected items."""
    data = request.json
    item_ids = data.get('items', [])

    if not item_ids:
        return jsonify({'error': 'No items selected'}), 400

    conn = get_db()

    placeholders = ','.join('?' * len(item_ids))
    rows = conn.execute(
        f"SELECT * FROM items WHERE id IN ({placeholders}) AND status = 'pending'",
        item_ids
    ).fetchall()

    if not rows:
        conn.close()
        return jsonify({'error': 'No pending items found with given IDs'}), 400

    items = []
    order_ids = set()
    for row in rows:
        qty_override = data.get('quantities', {}).get(str(row['id']))
        qty = int(qty_override) if qty_override else row['qty']
        items.append({
            'title': row['title'],
            'price': row['price'],
            'qty': qty,
            'order_id': row['order_id'],
            'tracking': row['tracking'] if 'tracking' in row.keys() else '',
        })
        order_ids.add(row['order_id'])

    # Apply discount if provided (silently — just use the reduced price)
    discount = float(data.get('discount', 0) or 0)
    if discount > 0:
        for item in items:
            item['price'] = round(item['price'] * (1 - discount / 100), 1)

    invoice_number = get_next_invoice_number()
    invoice_date = date.today()
    total = sum(i['price'] * i['qty'] for i in items)

    # Create invoice record (same table)
    cursor = conn.execute(
        "INSERT INTO invoices (invoice_number, date, total, items_count, type) VALUES (?, ?, ?, ?, ?)",
        (invoice_number, invoice_date.isoformat(), total, len(items), 'remito')
    )
    invoice_id = cursor.lastrowid

    # Update items status
    conn.execute(
        f"UPDATE items SET status = 'invoiced', invoice_id = ? WHERE id IN ({placeholders})",
        [invoice_id] + item_ids
    )
    conn.commit()

    # Generate remito PDFs
    pdf_buf = generate_remito_pdf(invoice_number, invoice_date, items)
    pdf_no_prices_buf = generate_remito_pdf_no_prices(invoice_number, invoice_date, items)

    # Generate TXT with trackings (same format as invoice)
    orders_grouped = defaultdict(list)
    for item in items:
        tracking = item.get('tracking', '') or item['order_id']
        orders_grouped[tracking].append(item['title'])
    lines = []
    for oid in sorted(orders_grouped.keys()):
        titles = orders_grouped[oid]
        if len(titles) > 1:
            truncated = [t[:60] for t in titles]
            lines.append(f"{oid} ({', '.join(truncated)})")
        else:
            lines.append(oid)
    txt_content = '\n'.join(lines)

    # Create ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'Remito_{invoice_number}.pdf', pdf_buf.read())
        zf.writestr(f'Remito_{invoice_number}-without-prices.pdf', pdf_no_prices_buf.read())
        zf.writestr(f'Remito_{invoice_number}_trackings.txt', txt_content)

    zip_buf.seek(0)
    zip_data = zip_buf.read()
    conn.close()

    # Try to send email if configured — individual files, no ZIP
    email_status = 'not_configured'
    email_address = ''
    email_error = ''
    cfg = load_config()
    to_email = cfg.get('email')
    if to_email:
        email_address = to_email
        try:
            subject = f"Invoice {invoice_number} — Zero"

            # Build body with trackings
            body_lines = [f"Remito #{invoice_number} — {invoice_date.strftime('%d/%m/%Y')}"]
            body_lines.append("")
            body_lines.append("Trackings:")
            body_lines.append("-" * 40)
            for line in lines:
                body_lines.append(line)
            body_lines.append("")
            body_lines.append(f"Total: ${total:,.2f}")
            body = '\n'.join(body_lines)

            # Regenerate buffers for email
            pdf_buf_email = generate_remito_pdf(invoice_number, invoice_date, items)
            pdf_np_email = generate_remito_pdf_no_prices(invoice_number, invoice_date, items)

            email_files = [
                (f'Remito_{invoice_number}.pdf', pdf_buf_email.read(), 'application/pdf'),
                (f'Remito_{invoice_number}-without-prices.pdf', pdf_np_email.read(), 'application/pdf'),
                (f'Remito_{invoice_number}_trackings.txt', txt_content.encode('utf-8'), 'text/plain'),
            ]
            send_email_with_attachments(email_files, to_email, subject, body)
            email_status = 'sent'
        except Exception as exc:
            email_status = 'error'
            email_error = str(exc)

    response = make_response(send_file(
        io.BytesIO(zip_data),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Remito_{invoice_number}.zip'
    ))
    response.headers['X-Email-Status'] = email_status
    response.headers['X-Email-Address'] = email_address
    response.headers['X-Email-Error'] = email_error
    response.headers['Access-Control-Expose-Headers'] = 'X-Email-Status, X-Email-Address, X-Email-Error'
    return response

@app.route('/api/remitos/<int:invoice_id>/download')
def download_remito(invoice_id):
    """Re-download a remito."""
    conn = get_db()
    inv = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Invoice not found'}), 404

    rows = conn.execute("SELECT * FROM items WHERE invoice_id = ?", (invoice_id,)).fetchall()
    conn.close()

    items = [{'title': r['title'], 'price': r['price'], 'qty': r['qty'], 'order_id': r['order_id'], 'tracking': r['tracking'] if 'tracking' in r.keys() else ''} for r in rows]
    invoice_date = datetime.strptime(inv['date'], '%Y-%m-%d').date()

    pdf_buf = generate_remito_pdf(inv['invoice_number'], invoice_date, items)
    pdf_no_prices_buf = generate_remito_pdf_no_prices(inv['invoice_number'], invoice_date, items)

    # Generate TXT
    orders_grouped = defaultdict(list)
    for item in items:
        tracking = item.get('tracking', '') or item['order_id']
        orders_grouped[tracking].append(item['title'])
    lines = []
    for oid in sorted(orders_grouped.keys()):
        titles = orders_grouped[oid]
        if len(titles) > 1:
            truncated = [t[:60] for t in titles]
            lines.append(f"{oid} ({', '.join(truncated)})")
        else:
            lines.append(oid)
    txt_content = '\n'.join(lines)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'Remito_{inv["invoice_number"]}.pdf', pdf_buf.read())
        zf.writestr(f'Remito_{inv["invoice_number"]}-without-prices.pdf', pdf_no_prices_buf.read())
        zf.writestr(f'Remito_{inv["invoice_number"]}_trackings.txt', txt_content)

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Remito_{inv["invoice_number"]}.zip'
    )

@app.route('/api/invoices')
def get_invoices():
    """Get invoice history."""
    conn = get_db()
    invoices = conn.execute(
        "SELECT * FROM invoices ORDER BY invoice_number DESC"
    ).fetchall()

    result = []
    for inv in invoices:
        items = conn.execute(
            "SELECT * FROM items WHERE invoice_id = ?", (inv['id'],)
        ).fetchall()

        inv_type = 'invoice'
        try:
            inv_type = inv['type'] or 'invoice'
        except (IndexError, KeyError):
            pass

        result.append({
            'id': inv['id'],
            'invoice_number': inv['invoice_number'],
            'date': inv['date'],
            'total': inv['total'],
            'items_count': inv['items_count'],
            'created_at': inv['created_at'],
            'type': inv_type,
            'items': [dict(i) for i in items],
            'all_sent': all(i['status'] == 'sent' for i in items),
        })

    conn.close()
    return jsonify(result)

@app.route('/api/invoices/<int:invoice_id>/download')
def download_invoice(invoice_id):
    """Re-download an invoice."""
    conn = get_db()
    inv = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Invoice not found'}), 404

    rows = conn.execute("SELECT * FROM items WHERE invoice_id = ?", (invoice_id,)).fetchall()
    conn.close()

    items = [{'title': r['title'], 'price': r['price'], 'qty': r['qty'], 'order_id': r['order_id'], 'tracking': r['tracking'] if 'tracking' in r.keys() else ''} for r in rows]

    invoice_date = datetime.strptime(inv['date'], '%Y-%m-%d').date()

    xlsx_buf = generate_xlsx(inv['invoice_number'], invoice_date, items)
    pdf_buf = generate_pdf(inv['invoice_number'], invoice_date, items)
    remito_no_prices_buf = generate_remito_pdf_no_prices(inv['invoice_number'], invoice_date, items)

    # Generate TXT with trackings (grouped with item names)
    orders_grouped = defaultdict(list)
    for item in items:
        tracking = item.get('tracking', '') or item['order_id']
        orders_grouped[tracking].append(item['title'])
    lines = []
    for oid in sorted(orders_grouped.keys()):
        titles = orders_grouped[oid]
        if len(titles) > 1:
            truncated = [t[:60] for t in titles]
            lines.append(f"{oid} ({', '.join(truncated)})")
        else:
            lines.append(oid)
    txt_content = '\n'.join(lines)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'Invoice_{inv["invoice_number"]}.xlsx', xlsx_buf.read())
        zf.writestr(f'Invoice_{inv["invoice_number"]}.pdf', pdf_buf.read())
        zf.writestr(f'Remito_{inv["invoice_number"]}-sin-precios.pdf', remito_no_prices_buf.read())
        zf.writestr(f'Invoice_{inv["invoice_number"]}_trackings.txt', txt_content)

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Invoice_{inv["invoice_number"]}.zip'
    )

@app.route('/api/invoices/<int:invoice_id>/mark-sent', methods=['POST'])
def mark_sent(invoice_id):
    """Mark all items in an invoice as sent."""
    conn = get_db()
    conn.execute("UPDATE items SET status = 'sent' WHERE invoice_id = ?", (invoice_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Invoice marked as sent'})

@app.route('/api/items/<int:item_id>/status', methods=['PUT'])
def update_item_status(item_id):
    """Change the status of an item (pending/invoiced/sent)."""
    data = request.json
    new_status = data.get('status')
    if new_status not in ('pending', 'invoiced', 'sent'):
        return jsonify({'error': 'Invalid status'}), 400
    conn = get_db()
    # If changing to pending, unlink from invoice
    if new_status == 'pending':
        conn.execute("UPDATE items SET status = ?, invoice_id = NULL WHERE id = ?", (new_status, item_id))
    else:
        conn.execute("UPDATE items SET status = ? WHERE id = ?", (new_status, item_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Status updated'})

@app.route('/api/invoices/<int:invoice_id>', methods=['DELETE'])
def delete_invoice(invoice_id):
    """Delete an invoice and return its items to pending."""
    conn = get_db()
    inv = conn.execute("SELECT id FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if inv is None:
        conn.close()
        return jsonify({'error': 'Invoice no encontrado'}), 404
    # Return items back to pending status
    conn.execute("UPDATE items SET status = 'pending', invoice_id = NULL WHERE invoice_id = ?", (invoice_id,))
    # Delete invoice record
    conn.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Invoice eliminado, items devueltos a pendientes'})

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    """Delete an item (any status)."""
    conn = get_db()
    if conn.execute("SELECT id FROM items WHERE id = ?", (item_id,)).fetchone() is None:
        conn.close()
        return jsonify({'error': 'Item no encontrado'}), 404
    conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Item deleted'})

# ─── Config API ─────────────────────────────────────────────────────────────

@app.route('/api/config/po', methods=['GET'])
def get_po_config():
    return jsonify({'pos': get_accepted_pos()})

@app.route('/api/config/invoice-counter', methods=['GET'])
def get_invoice_counter():
    conn = get_db()
    row = conn.execute("SELECT MAX(invoice_number) as max_num FROM invoices").fetchone()
    conn.close()
    last = row['max_num'] if row and row['max_num'] else 2069
    # Also check config override
    cfg = load_config()
    override = cfg.get('last_invoice_number')
    if override and override > last:
        last = override
    return jsonify({'last_number': last})

@app.route('/api/config/invoice-counter', methods=['PUT'])
def set_invoice_counter():
    data = request.get_json()
    num = data.get('last_number')
    if not num or not isinstance(num, int) or num < 1:
        return jsonify({'error': 'Número inválido'}), 400
    cfg = load_config()
    cfg['last_invoice_number'] = num
    save_config(cfg)
    return jsonify({'last_number': num})

@app.route('/api/config/po', methods=['PUT'])
def set_po_config():
    data = request.get_json()
    pos = data.get('pos', [])
    if not pos or not isinstance(pos, list):
        return jsonify({'error': 'Debe enviar una lista de POs'}), 400
    # Clean and deduplicate
    pos = list(dict.fromkeys(str(p).strip() for p in pos if str(p).strip()))
    cfg = load_config()
    cfg['accepted_pos'] = pos
    save_config(cfg)
    return jsonify({'pos': pos})

@app.route('/api/config/email', methods=['GET'])
def get_email_config():
    cfg = load_config()
    return jsonify({'email': cfg.get('email', None)})

@app.route('/api/config/email', methods=['PUT'])
def set_email_config():
    data = request.get_json()
    email = data.get('email', None)
    if email is not None and email != '' and '@' not in str(email):
        return jsonify({'error': 'Email inválido'}), 400
    cfg = load_config()
    cfg['email'] = email if email else None
    save_config(cfg)
    return jsonify({'email': cfg['email']})

@app.route('/api/config/smtp', methods=['GET'])
def get_smtp_config():
    cfg = load_config()
    smtp = cfg.get('smtp', {})
    return jsonify({
        'server': smtp.get('server', 'smtp.gmail.com'),
        'port': smtp.get('port', 587),
        'user': smtp.get('user', ''),
        'password_set': bool(smtp.get('password', '')),
    })

@app.route('/api/config/smtp', methods=['PUT'])
def set_smtp_config():
    data = request.get_json()
    cfg = load_config()
    smtp = cfg.get('smtp', {})
    if 'server' in data:
        smtp['server'] = str(data['server']).strip()
    if 'port' in data:
        smtp['port'] = int(data['port'])
    if 'user' in data:
        smtp['user'] = str(data['user']).strip()
    if 'password' in data and data['password']:
        smtp['password'] = str(data['password'])
    cfg['smtp'] = smtp
    save_config(cfg)
    return jsonify({'ok': True, 'server': smtp.get('server'), 'port': smtp.get('port'), 'user': smtp.get('user'), 'password_set': bool(smtp.get('password', ''))})

@app.route('/api/config/smtp/test', methods=['POST'])
def test_smtp_config():
    cfg = load_config()
    to_email = cfg.get('email') or (request.get_json() or {}).get('to')
    if not to_email:
        return jsonify({'ok': False, 'error': 'No hay email destino configurado'}), 400
    try:
        send_email_with_attachments(
            [], to_email,
            subject='Test Email — Invoice Builder',
            body='Este es un email de prueba enviado desde Invoice Builder.\n\nSi lo recibiste, el envío de emails está funcionando correctamente.'
        )
        return jsonify({'ok': True, 'sent_to': to_email})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 500

@app.route('/api/config/resend', methods=['GET'])
def get_resend_config():
    cfg = load_config()
    return jsonify({
        'api_key_set': bool(cfg.get('resend_api_key', '')),
        'from_email': cfg.get('resend_from', ''),
    })

@app.route('/api/config/resend', methods=['PUT'])
def set_resend_config():
    data = request.get_json()
    cfg = load_config()
    if 'api_key' in data and data['api_key']:
        cfg['resend_api_key'] = str(data['api_key']).strip()
    if 'from_email' in data:
        cfg['resend_from'] = str(data['from_email']).strip()
    save_config(cfg)
    return jsonify({'ok': True, 'api_key_set': bool(cfg.get('resend_api_key', '')), 'from_email': cfg.get('resend_from', '')})

@app.route('/sw.js')
def service_worker():
    return send_file(os.path.join(BASE_DIR, 'static', 'sw.js'), mimetype='application/javascript')

# ─── Main ───────────────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
